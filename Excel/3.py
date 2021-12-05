import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border 
from openpyxl.styles import Side, PatternFill, Color
from openpyxl.styles import GradientFill, colors, NamedStyle
loc = 'C:/Users/allen/Downloads/Resources/' \
      'Python/Lib/L1/Excel/data/2'                              # Location of the directory
wb = Workbook()                                                 # Create workbook
ws = wb.active                                                  # Go to active sheet of wb

for z in range(1,20): ws.append(range(300))                     # Add data to sheet

ws.merge_cells('B2:F7')                                         # Merge cells B2:F7
ws.unmerge_cells('B2:F7')                                       # Unmerge cells B2:F7
ws.merge_cells(
    start_row=2, end_row=7,
    start_column=2, end_column=6
)                                                               # Merge cells B2:F7 

# Merged Cell Formatting
cell = ws['B2']                                                 # Get cell B2
cell.value = 'Allen'                                            # Set value
cell.font = Font(
    name='Calibri', color=colors.COLOR_INDEX[44],
    size=28, bold=True
)                                                               # Set font
cell.alignment = Alignment(
    horizontal='center', vertical='center'
)                                                               # Set alignment
cell.fill = GradientFill(
    stop=('009FFD', '2A2A72'), type='linear'
)                                                               # Set gradient fill

# Highlight Style 1
h1 = NamedStyle(name='highlight1')                              # Create a style
h1.font = Font(bold=True)                                       # Set font
bd1 = Side(style='thin', color='FFFF00')                        # Set border
h1.fill = PatternFill('solid', fgColor='FFFF00')                # Set fill
h1.border = Border(
    left=bd1, top=bd1, right=bd1, bottom=bd1
)                                                               # Apply border

# Highlight Style 2
h2 = NamedStyle(name='highlight2')                              # Create a style
h2.font = Font(
    color=colors.COLOR_INDEX[1],
    bold='True', underline='single')                            # Set font
bd2 = Side(style='thick', color='000000')                       # Set border
h2.fill = GradientFill(
    stop=('333252', '030712'), type='linear'
)                                                               # Set fill
h2.border = Border(
    left=bd2, top=bd2, right=bd2, bottom=bd2
)                                                               # Apply border

# Apply highlights
z = 0 
for col in ws.iter_cols(
    min_row=1, max_row=8, 
    min_col=8, max_col=11
):                                                              # Iterate through columns
    col[z+1].style = h1                                         # Apply highlight1
    col[z+3].style = h2                                         # Apply highlight2
    z+=1

wb.save(f'{loc}/B_Write.xlsx')                                  # Save workbook