import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
loc = 'C:/Users/allen/Downloads/Resources/' \
      'Python/Lib/L1/Excel/data/2'                              # Location of the directory

# Create   
wb1 = Workbook()                                                # Create workbook
wb2 = load_workbook(f'{loc}/A_Read.xlsx')                       # Load workbook
wb3 = load_workbook(f'{loc}/A_Read.xlsx')                       # Load workbook
ws1 = wb1.active                                                # Go to active sheet of wb1
ws2 = wb2.active                                                # Go to active sheet of wb2
ws3 = wb2.active                                                # Go to active sheet of wb3

# Label
ws1.title = 'Christi'                                           # Title of active sheet
s1 = wb1.create_sheet('Allen')                                  # Create sheet Allen
s2 = wb1.create_sheet('Alvin', 0)                               # Create sheet Alvin at index 0
print(wb1.sheetnames)                                           # Print all sheet names

# Modify
c2 = ws2['A1']                                                  # Select A1 of active sheet
print(c2.value)                                                 # Prints cell value
ws2['A1'] = 'Direction'                                         # Modify cell value
wb2.save(f'{loc}/A_Write.xlsx')                                 # Save in new file

# Display
print(ws3['A1':'C1'],'\n\n')                                    # 1D tuple from cells A1 to C1
print(ws3['A':'C'],'\n\n')                                      # 2D tuple from columns A to C
print(ws3[1:5],'\n\n')                                          # 2D tuple from rows 1 to 5
for z in ws3.iter_rows(
    min_row=1, max_row=2,
    min_col=1, max_col=4, 
    values_only=True                                            # False would return objects
):                                                              
    for y in z: print(y)                                        # Values of A1:D2