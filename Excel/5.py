import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
loc = 'C:/Users/allen/Downloads/Resources/' \
      'Python/Lib/L1/Excel/data/3'                                # Location of the directory

# Merge Dataframes
df1 = pd.read_excel(f'{loc}/A_Read.xlsx', sheet_name='Sheet1')    # Read the first file sheet 1
df2 = pd.read_excel(f'{loc}/A_Read.xlsx', sheet_name='Sheet2')    # Read the first file sheet 2        
df3 = pd.read_excel(f'{loc}/B_Read.xlsx')                         # Read the second file
df = pd.concat([df1, df2, df3], sort=False)                       # Combine the dataframes
df.to_excel(f'{loc}/D_Write.xlsx')                                # Write the dataframe to a file

# Modify Workbooks
wb = load_workbook(f'{loc}/D_Write.xlsx')                         # Load the file
ws = wb.active                                                    # Get the active sheet
ws['H1'].font = Font(bold=True)                                   # Set the font to bold
ws['H1'].value = 'Total'                                          # Set the value

# New Column using Workbook
for z in range(2,300):                                            # Loop through the rows
      f = ws['F'+str(z)].value                                    # Value of cell in E column
      g = ws['G'+str(z)].value                                    # Value of cell in F column
      ws['H'+str(z)].value = f + g                                # Save the calculated value

# New Column using Dataframe
df['Total'] = df['Cost per'] + df['Units Sold']                   # Calculate the total
rows = dataframe_to_rows(df, index=False, header=True)            # Convert the dataframe to rows
for z,a in enumerate(rows,1): 
      for y,b in enumerate(a,2):                                   
            ws.cell(row=z, column=y).value = b                    # Save the calculated value

wb.save(f'{loc}/D_Write.xlsx')                                    # Save the file